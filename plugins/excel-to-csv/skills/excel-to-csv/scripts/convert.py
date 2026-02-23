#!/usr/bin/env python3
"""
<<<<<<< HEAD
excel_to_csv.py (CLI)
=======
convert_master_sheet_to_csv.py (CLI)
>>>>>>> origin/main
=====================================

Purpose:
    Convert all (or selected) sheets from an Excel workbook into CSV files.

<<<<<<< HEAD
Layer: Data Processing Utilities

Usage Examples:
    python3 scripts/convert.py --excel data.xlsx --sheets "Sheet1" --outdir ./output
    python3 scripts/convert.py --excel data.xlsx --sheets "SalesTable" --outdir ./output

Supported Object Types:
    - .xlsx
    - .xls

CLI Arguments:
    --excel         : Path to the Excel workbook.
    --outdir        : Output directory for CSV files (default: current directory).
    --sheets        : Comma-separated list of sheet names to convert.
    --write-empty   : Write a CSV even if the sheet has no non-empty rows/columns.

Output:
    - CSV files generated in the specified output directory.

Key Functions:
    - sanitize_sheet_name()
    - convert_excel_to_csv()
"""

=======
Layer: Curate / Utilities

Usage Examples:
    python tools/curate/utils/convert_master_sheet_to_csv.py --help

Supported Object Types:
    - Generic

CLI Arguments:
    --excel         : Path to the Excel workbook (.xlsx). If omitted, the default JAM master path in this folder will be used.
    --outdir        : Output directory for CSV files (default: ./csv)
    --sheets        : Comma-separated list of sheet names to convert (default: all sheets)
    --write-empty   : Write a CSV even if the sheet has no non-empty rows/columns
    --same-name     : When converting a workbook with exactly one non-empty sheet, also write the CSV with the same base name as the workbook (e.g., JAM-SharePoint-MasterSheet.csv)

Input Files:
    - (See code)

Output:
    - (See code)

Key Functions:
    - sanitize_sheet_name(): No description.
    - convert_excel_to_csv(): No description.
    - main(): No description.

Script Dependencies:
    (None detected)

Consumed by:
    (Unknown)
"""
>>>>>>> origin/main
from pathlib import Path
import argparse
import sys
import re

try:
    import pandas as pd
<<<<<<< HEAD
    import openpyxl
    from openpyxl.utils.cell import range_boundaries
except ImportError as e:
    print("Missing dependency: pandas or openpyxl. Install with `pip install pandas openpyxl` and try again.", file=sys.stderr)
    sys.exit(1)


def sanitize_sheet_name(name: str) -> str:
    """Make filename-safe: replace spaces and illegal chars."""
    name = name.strip()
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = name.replace(" ", "_")
=======
except Exception as e:
    print("Missing dependency: pandas. Install with `pip install pandas openpyxl` and try again.")
    raise e


def sanitize_sheet_name(name: str) -> str:
    # make filename-safe: replace spaces and illegal chars
    name = name.strip()
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = name.replace(" ", "_")
    # truncate if very long
>>>>>>> origin/main
    return name[:120]


def convert_excel_to_csv(excel_path: Path, out_dir: Path, sheets=None, write_empty=False, encoding="utf-8-sig") -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
<<<<<<< HEAD
    
    target_names = [s.strip() for s in sheets.split(",")] if sheets else None
    
    # 1. Inspect workbook for sheets and tables using openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Maps requested name -> {'type': 'sheet'|'table', 'sheet_name': str, 'bounds': (min_col, min_row, max_col, max_row)}
    extractions = {}
    
    if target_names:
        for target in target_names:
            found = False
            # Check if it's a sheet
            if target in wb.sheetnames:
                extractions[target] = {'type': 'sheet', 'sheet_name': target}
                found = True
            else:
                # Check if it's a table in any sheet
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    if target in ws.tables:
                        tbl = ws.tables[target]
                        bounds = range_boundaries(tbl.ref)
                        extractions[target] = {'type': 'table', 'sheet_name': sheet, 'bounds': bounds}
                        found = True
                        break
            if not found:
                print(f"Warning: '{target}' not found as a sheet or table. Skipping.", file=sys.stderr)
    else:
        # Default behavior: extract all sheets
        for sheet in wb.sheetnames:
            extractions[sheet] = {'type': 'sheet', 'sheet_name': sheet}

    # 2. Extract parsed targets using pandas
    summary = {"written": [], "skipped": []}
    
    # Cache loaded dataframes per sheet so we only read them once
    loaded_sheets = {}

    for target, info in extractions.items():
        sheet_name = info['sheet_name']
        
        if sheet_name not in loaded_sheets:
            try:
                # Always load without headers so slicing by exact row indices matches openpyxl (1-indexed mapping)
                df = pd.read_excel(excel_path, sheet_name=sheet_name, engine="openpyxl", header=None)
                loaded_sheets[sheet_name] = df
            except Exception as e:
                print(f"Error reading sheet '{sheet_name}' from {excel_path}: {e}", file=sys.stderr)
                summary["skipped"].append(target)
                continue
                
        df = loaded_sheets[sheet_name]
        
        if df is None or df.empty:
            if write_empty:
                out_path = out_dir / f"{sanitize_sheet_name(target)}.csv"
                out_path.write_text("", encoding=encoding)
                summary["written"].append(str(out_path))
            else:
                summary["skipped"].append(target)
            continue
            
        # If it's a table, slice the dataframe
        extracted_df = df
        if info['type'] == 'table':
            min_col, min_row, max_col, max_row = info['bounds']
            # Pandas is 0-indexed, openpyxl is 1-indexed
            extracted_df = df.iloc[min_row-1:max_row, min_col-1:max_col]
        
        # Drop rows & columns that are completely empty first
        extracted_df = extracted_df.dropna(axis=0, how="all").dropna(axis=1, how="all")

        # Determine headers for the slice
        if not extracted_df.empty and len(extracted_df) > 0:
            extracted_df = extracted_df.copy()
            
            # If extracting a sheet, try to find the actual header row (first row with multiple distinct values)
            # to handle cases where a title string extends into empty leading columns
            first_row_idx = extracted_df.index[0]
            if info['type'] == 'sheet' and len(extracted_df) > 1:
                for idx, row in extracted_df.iterrows():
                    # If a row has more than 1 non-null value, consider it the real header
                    if row.notna().sum() > 1:
                        first_row_idx = idx
                        # Drop columns that are entirely empty AFTER this header row
                        extracted_df = extracted_df.loc[first_row_idx:].dropna(axis=1, how="all")
                        break

            extracted_df.columns = extracted_df.loc[first_row_idx].fillna(f"Unnamed")
            # Drop that header row (and any preceding junk rows) from the data body
            extracted_df = extracted_df.loc[first_row_idx + 1:]

        df2 = extracted_df

        if df2.empty and not write_empty:
            summary["skipped"].append(target)
            continue

        out_path = out_dir / f"{sanitize_sheet_name(target)}.csv"
        df2.to_csv(out_path, index=False, encoding=encoding)
        summary["written"].append(str(out_path))
=======

    # Read either all sheets or the selected ones
    if sheets:
        # Accept comma-separated list
        sheet_list = [s.strip() for s in sheets.split(",") if s.strip()]
        # pandas accepts list of names
        df_map = pd.read_excel(excel_path, sheet_name=sheet_list, engine="openpyxl")
    else:
        df_map = pd.read_excel(excel_path, sheet_name=None, engine="openpyxl")

    summary = {"written": [], "skipped": []}

    last_written = None
    written_count = 0
    for sheet_name, df in df_map.items():
        # If df is None (happens if sheet empty) create empty DataFrame
        if df is None:
            if write_empty:
                out_path = out_dir / f"{sanitize_sheet_name(sheet_name)}.csv"
                out_path.write_text("", encoding=encoding)
                summary["written"].append(str(out_path))
                last_written = (sheet_name, df)
                written_count += 1
            else:
                summary["skipped"].append(sheet_name)
            continue

        # Drop rows & columns that are completely empty
        df2 = df.dropna(axis=0, how="all")
        df2 = df2.dropna(axis=1, how="all")

        if df2.empty and not write_empty:
            summary["skipped"].append(sheet_name)
            continue

        out_path = out_dir / f"{sanitize_sheet_name(sheet_name)}.csv"
        df2.to_csv(out_path, index=False, encoding=encoding)
        summary["written"].append(str(out_path))
        last_written = (sheet_name, df2)
        written_count += 1

    # If --same-name requested and exactly one non-empty/written sheet, also write a CSV named after the workbook base
    if written_count == 1 and 'same_name' in locals() and locals().get('same_name'):
        # locals usage is a precaution — we set this in the caller via args
        base_name = excel_path.stem
        single_out = out_dir / f"{base_name}.csv"
        # last_written[1] is the dataframe (or None if it was empty-but-write_empty)
        if last_written and last_written[1] is not None:
            last_written[1].to_csv(single_out, index=False, encoding=encoding)
            summary["written"].append(str(single_out))
        else:
            # If the sheet was empty but write_empty==True, create an empty file
            single_out.write_text("", encoding=encoding)
            summary["written"].append(str(single_out))
>>>>>>> origin/main

    return summary


def main():
    parser = argparse.ArgumentParser(description="Convert Excel workbook sheets to CSV files")
<<<<<<< HEAD
    parser.add_argument("--excel", "-e", required=True, help="Path to the Excel workbook (.xlsx).")
    parser.add_argument("--outdir", "-o", default=".", help="Output directory for CSV files (default: .)")
    parser.add_argument("--sheets", "-s", default=None, help="Comma-separated list of sheet names to convert (default: all sheets)")
    parser.add_argument("--write-empty", action="store_true", help="Write a CSV even if the sheet is empty")

    args = parser.parse_args()
    excel_path = Path(args.excel)

    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    out_dir = Path(args.outdir)
    print(f"Reading: {excel_path}")
    print(f"Output directory: {out_dir}")

    result = convert_excel_to_csv(excel_path, out_dir, sheets=args.sheets, write_empty=args.write_empty)

    print("\nSummary:")
    print("  Written:")
    if not result["written"]:
        print("    (None)")
    for p in result["written"]:
        print(f"    {p}")
        
    print("  Skipped (empty):")
    if not result["skipped"]:
        print("    (None)")
    for s in result["skipped"]:
        print(f"    {s}")
=======
    parser.add_argument("--excel", "-e", required=True,
                        help="Path to the Excel workbook (.xlsx).")
    parser.add_argument("--outdir", "-o", default="./csv",
                        help="Output directory for CSV files (default: ./csv)")
    parser.add_argument("--sheets", "-s", default=None,
                        help="Comma-separated list of sheet names to convert (default: all sheets)")
    parser.add_argument("--write-empty", action="store_true",
                        help="Write a CSV even if the sheet has no non-empty rows/columns")
    parser.add_argument("--same-name", action="store_true",
                        help="When converting a workbook with exactly one non-empty sheet, also write the CSV with the same base name as the workbook")

    args = parser.parse_args()

    excel_path = Path(args.excel)

    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        sys.exit(2)

    # If --same-name requested, write outputs into the same folder as the Excel file
    if args.same_name:
        out_dir = excel_path.parent
    else:
        out_dir = Path(args.outdir)

    print(f"Reading: {excel_path}")
    print(f"Output directory: {out_dir}")

    try:
        result = convert_excel_to_csv(excel_path, out_dir, sheets=args.sheets, write_empty=args.write_empty)
    except Exception as exc:
        print("Error converting workbook:", exc)
        sys.exit(3)

    # If requested, and exactly one CSV written, also write a copy named after the workbook base
    if args.same_name and len(result.get("written", [])) == 1:
        import shutil
        src = Path(result["written"][0])
        dst = out_dir / f"{excel_path.stem}.csv"
        shutil.copyfile(src, dst)
        # Add to summary so it is listed below
        result["written"].append(str(dst))
        print(f"Also wrote single-sheet CSV with workbook base name: {dst}")

    print("Summary:")
    print("  Written:")
    for p in result["written"]:
        print("    ", p)
    print("  Skipped (empty):")
    for s in result["skipped"]:
        print("    ", s)
>>>>>>> origin/main


if __name__ == "__main__":
    main()
